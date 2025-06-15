import csv
import os
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
from reportlab.lib.units import inch

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


class ProductProcessor:
    def __init__(self, db_manager):
        self.db_manager = db_manager

    def add_new_product(self, user_id, product_data_dict):
        if not user_id:
            return False, "User ID missing.", None
        if not product_data_dict.get('product_name') or product_data_dict.get('selling_price') is None:
            return False, "Product name and selling price are required.", None
        
        try:
            product_data_dict['selling_price'] = float(product_data_dict.get('selling_price', 0.0))
            for field in ['purchase_price', 'stock_quantity', 'low_stock_threshold']:
                value = product_data_dict.get(field)
                if value is not None and str(value).strip() != "":
                    if 'price' in field:
                        product_data_dict[field] = float(value)
                    else:
                        product_data_dict[field] = int(value)
        except (ValueError, TypeError):
            return False, "Invalid numeric value for price or quantity.", None

        product_id, message = self.db_manager.add_product(user_id, product_data_dict)
        return product_id is not None, message, product_id

    def get_products_for_display(self, user_id, **kwargs):
        if not user_id: return []
        products = self.db_manager.get_products_by_user_id(user_id, **kwargs)
        return products

    def get_single_product_details(self, user_id, product_id):
        if not user_id or not product_id: return None
        return self.db_manager.get_product_by_id_and_user_id(product_id, user_id)

    def update_product_details(self, user_id, product_id, product_data_dict):
        if not user_id or not product_id:
            return False, "User or Product ID missing for update."
        
        return self.db_manager.update_product(product_id, user_id, product_data_dict)

    def remove_product(self, user_id, product_id):
        if not user_id or not product_id:
            return False, "User or Product ID missing for delete."

        return self.db_manager.delete_product(product_id, user_id)

    def adjust_product_stock(self, user_id, product_id, adjustment_value):
        """
        Adjusts the stock for a single product.
        'adjustment_value' can be positive (to add stock) or negative (to remove stock).
        """
        if not all([user_id, product_id]):
            return False, "User or Product ID missing."
        
        if adjustment_value == 0:
            return True, "No adjustment made." 

        product = self.db_manager.get_product_by_id_and_user_id(product_id, user_id)
        if not product:
            return False, "Product not found."
            
        current_stock = product['stock_quantity']
        new_stock = current_stock + adjustment_value

        if new_stock < 0:
            return False, f"Cannot adjust stock by {adjustment_value}. It would result in a negative quantity ({new_stock})."

        update_data = {'stock_quantity': new_stock}
        return self.db_manager.update_product(product_id, user_id, update_data)
    
    def export_products_and_sales_to_csv(self, user_id, file_path):
        """Exports all of a user's products and their associated sales to a CSV file."""
        try:
            sales_records = self.db_manager.get_sales_records(user_id=user_id)
            
            with open(file_path, 'w', newline='', encoding='utf-8') as csvfile:
                if not sales_records:
                    csvfile.write("No sales data to export.")
                    return True, "Export complete. No sales data found."
                
                header = sales_records[0].keys()
                writer = csv.DictWriter(csvfile, fieldnames=header)
                
                writer.writeheader()
                writer.writerows(sales_records)
                
            return True, f"Data successfully exported to {os.path.basename(file_path)}"
        except Exception as e:
            return False, f"An error occurred during export: {e}"
        

    def export_products_and_sales_to_pdf(self, user_id, file_path):
        """Exports all of a user's products and their sales to a PDF file."""
        try:
            doc = SimpleDocTemplate(file_path)
            story = []
            styles = getSampleStyleSheet()

            title = Paragraph("Product and Sales Report", styles['h1'])
            story.append(title)
            story.append(Spacer(1, 0.2*inch))

            sales_records = self.db_manager.get_sales_records(user_id=user_id)
            if not sales_records:
                story.append(Paragraph("No sales data to export.", styles['Normal']))
                doc.build(story)
                return True, "Export complete. No sales data found."
            
            header = ["Date", "Product Name", "Qty", "Unit Price", "Total"]
            data = [header]
            for record in sales_records:
                data.append([
                    record['sale_date'].split(" ")[0],
                    record['product_name'],
                    record['quantity_sold'],
                    f"${record['price_at_sale']:.2f}",
                    f"${record['total_revenue']:.2f}"
                ])

            table = Table(data, colWidths=[1.5*inch, 2.5*inch, 0.5*inch, 1*inch, 1*inch])
            style = TableStyle([
                ('BACKGROUND', (0,0), (-1,0), colors.grey),
                ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
                ('ALIGN', (0,0), (-1,-1), 'CENTER'),
                ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
                ('BOTTOMPADDING', (0,0), (-1,0), 12),
                ('BACKGROUND', (0,1), (-1,-1), colors.beige),
                ('GRID', (0,0), (-1,-1), 1, colors.black)
            ])
            table.setStyle(style)
            
            story.append(table)
            doc.build(story)
            
            return True, f"Data successfully exported to {os.path.basename(file_path)}"
        except Exception as e:
            return False, f"An error occurred during PDF export: {e}"
        
    
    def export_data_to_excel(self, user_id, file_path):
        """Exporte les produits et leurs ventes dans un fichier Excel (.xlsx) bien formaté."""
        try:
            sales_records = self.db_manager.get_sales_records(user_id=user_id)
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Sales Report"

            if not sales_records:
                ws['A1'] = "Aucune donnée de vente à exporter."
                wb.save(file_path)
                return True, "Export terminé. Aucune donnée de vente trouvée."

            headers = ["Date", "Product Name", "SKU", "Category", "Quantity Sold", "Price at Sale", "Total Revenue"]
            ws.append(headers)

            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4F4F4F", end_color="4F4F4F", fill_type="solid")

            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill

            for record in sales_records:
                row_data = [
                    record['sale_date'],
                    record['product_name'],
                    record.get('sku', 'N/A'),
                    record.get('category', 'N/A'),
                    record['quantity_sold'],
                    record['price_at_sale'],
                    record['total_revenue']
                ]
                ws.append(row_data)

            for col_idx, column_cells in enumerate(ws.columns):
                max_length = 0
                column_letter = get_column_letter(col_idx + 1)
                for cell in column_cells:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column_letter].width = adjusted_width

            wb.save(file_path)
            return True, f"Données exportées avec succès vers {os.path.basename(file_path)}"

        except Exception as e:
            return False, f"Une erreur est survenue lors de l'export Excel : {e}"